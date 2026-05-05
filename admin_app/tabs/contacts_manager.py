"""Tab para gestionar contactos."""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget,
    QTableWidgetItem, QLabel, QLineEdit, QMessageBox, QFileDialog,
    QHeaderView, QProgressBar, QInputDialog, QTextEdit, QDialog,
    QFormLayout, QComboBox, QAbstractItemView,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal

from api_client import get_client


# ─── Threads ─────────────────────────────────────────────────────────────────

class ContactsLoaderThread(QThread):
    progress = pyqtSignal(str)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def run(self):
        self.progress.emit("Cargando contactos...")
        client = get_client()
        ok, contacts, err = client.get_all_contacts()
        if ok:
            self.finished.emit(contacts)
        else:
            self.error.emit(err)


class LabelsLoaderThread(QThread):
    finished = pyqtSignal(list)

    def run(self):
        client = get_client()
        ok, labels, _ = client.get_labels()
        self.finished.emit(labels if ok else [])


# ─── Diálogos ─────────────────────────────────────────────────────────────────

class ContactEditDialog(QDialog):
    """Diálogo para editar un contacto individualmente."""

    def __init__(self, contact: dict, parent=None):
        super().__init__(parent)
        self.contact = contact
        self.setWindowTitle("Editar Contacto")
        self.setMinimumWidth(420)
        self._init_ui()

    def _init_ui(self):
        layout = QFormLayout()

        phone_label = QLabel(
            self.contact.get("telefono") or self.contact.get("whatsapp_number", "N/A")
        )
        phone_label.setStyleSheet("font-weight: bold; color: #1565C0;")
        layout.addRow("Teléfono:", phone_label)

        self.nombre_input = QLineEdit(self.contact.get("nombre", ""))
        layout.addRow("Nombre:", self.nombre_input)

        self.email_input = QLineEdit(
            self.contact.get("correo") or self.contact.get("email", "")
        )
        layout.addRow("Email:", self.email_input)

        self.etiqueta_input = QLineEdit(self.contact.get("etiqueta_cliente", ""))
        layout.addRow("Etiqueta:", self.etiqueta_input)

        btns = QHBoxLayout()
        save_btn = QPushButton("💾 Guardar")
        save_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 5px 14px;")
        cancel_btn = QPushButton("Cancelar")
        save_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        btns.addWidget(save_btn)
        btns.addWidget(cancel_btn)
        layout.addRow(btns)

        self.setLayout(layout)

    def get_data(self) -> dict:
        return {
            "nombre": self.nombre_input.text().strip(),
            "correo": self.email_input.text().strip(),
            "etiqueta_cliente": self.etiqueta_input.text().strip(),
        }


class SendMessageDialog(QDialog):
    """Diálogo para componer y enviar mensaje a un grupo de contactos."""

    def __init__(self, phones: list, parent=None):
        super().__init__(parent)
        self.phones = phones
        self.setWindowTitle(f"Enviar mensaje a {len(phones)} contacto(s)")
        self.setMinimumWidth(500)
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout()

        info = QLabel(f"📩 Se enviará a <b>{len(self.phones)}</b> contacto(s)")
        info.setStyleSheet("font-size: 13px; margin-bottom: 6px;")
        layout.addWidget(info)

        layout.addWidget(QLabel("Mensaje:"))
        self.message_input = QTextEdit()
        self.message_input.setPlaceholderText("Escribe el mensaje...")
        self.message_input.setMinimumHeight(130)
        layout.addWidget(self.message_input)

        btns = QHBoxLayout()
        send_btn = QPushButton("✉️ Enviar ahora")
        send_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 6px 18px;")
        cancel_btn = QPushButton("Cancelar")
        send_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        btns.addWidget(send_btn)
        btns.addWidget(cancel_btn)
        layout.addLayout(btns)

        self.setLayout(layout)

    def get_message(self) -> str:
        return self.message_input.toPlainText().strip()


# ─── Tab principal ─────────────────────────────────────────────────────────────

class ContactsManagerTab(QWidget):
    """Pestaña para gestionar contactos."""

    def __init__(self):
        super().__init__()
        self.contacts: list = []
        self.filtered_contacts: list = []
        self.has_loaded_once = False
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout()

        # ── Filtros ──────────────────────────────────────────────────────────
        filter_row = QHBoxLayout()

        filter_row.addWidget(QLabel("🔍"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Buscar por nombre, teléfono o email…")
        self.search_input.textChanged.connect(self._apply_filters)
        filter_row.addWidget(self.search_input)

        filter_row.addWidget(QLabel("🏷️ Etiqueta:"))
        self.label_filter = QComboBox()
        self.label_filter.addItem("Todas")
        self.label_filter.setMinimumWidth(160)
        self.label_filter.currentTextChanged.connect(self._apply_filters)
        filter_row.addWidget(self.label_filter)

        reload_labels_btn = QPushButton("↻")
        reload_labels_btn.setToolTip("Recargar etiquetas")
        reload_labels_btn.setFixedWidth(32)
        reload_labels_btn.clicked.connect(self._load_labels)
        filter_row.addWidget(reload_labels_btn)

        layout.addLayout(filter_row)

        # ── Tabla ─────────────────────────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(
            ["✓", "Teléfono", "Nombre", "Email", "Etiqueta", "Fecha"]
        )
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 36)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.doubleClicked.connect(self._on_double_click)
        layout.addWidget(self.table)

        self.count_label = QLabel("Sin datos. Presiona 'Recargar'.")
        layout.addWidget(self.count_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

        # ── Botones de acción principal ───────────────────────────────────────
        action_row = QHBoxLayout()

        btn_reload = QPushButton("🔄 Recargar")
        btn_reload.clicked.connect(self.load_contacts)

        btn_sel_all = QPushButton("☑ Selec. todos")
        btn_sel_all.clicked.connect(self._select_all)

        btn_desel = QPushButton("☐ Deselec.")
        btn_desel.clicked.connect(self._deselect_all)

        btn_send_sel = QPushButton("✉️ Enviar a seleccionados")
        btn_send_sel.setStyleSheet("background-color: #1976D2; color: white;")
        btn_send_sel.clicked.connect(self._send_to_selected)

        btn_send_all = QPushButton("📣 Enviar a todos (visibles)")
        btn_send_all.setStyleSheet("background-color: #6A1B9A; color: white;")
        btn_send_all.clicked.connect(self._send_to_all_visible)

        btn_edit = QPushButton("✏️ Editar")
        btn_edit.clicked.connect(self._edit_from_button)

        btn_delete = QPushButton("🗑️ Eliminar")
        btn_delete.setStyleSheet("background-color: #c62828; color: white;")
        btn_delete.clicked.connect(self._delete_contact)

        for b in [btn_reload, btn_sel_all, btn_desel, btn_send_sel,
                  btn_send_all, btn_edit, btn_delete]:
            action_row.addWidget(b)
        action_row.addStretch()
        layout.addLayout(action_row)

        # ── Botones I/O Excel ─────────────────────────────────────────────────
        io_row = QHBoxLayout()

        btn_exp = QPushButton("📥 Exportar Excel")
        btn_exp.clicked.connect(self._export_all)

        btn_exp_lbl = QPushButton("🏷️ Exportar por Etiqueta")
        btn_exp_lbl.clicked.connect(self._export_by_label)

        btn_exp_date = QPushButton("📅 Exportar por Fecha")
        btn_exp_date.clicked.connect(self._export_by_date)

        btn_tpl = QPushButton("📄 Descargar Plantilla")
        btn_tpl.clicked.connect(self._download_template)

        btn_imp = QPushButton("📤 Importar desde Excel")
        btn_imp.setStyleSheet("background-color: #E65100; color: white;")
        btn_imp.clicked.connect(self._import_contacts)

        for b in [btn_exp, btn_exp_lbl, btn_exp_date, btn_tpl, btn_imp]:
            io_row.addWidget(b)
        io_row.addStretch()
        layout.addLayout(io_row)

        # ── Snapshot ──────────────────────────────────────────────────────────
        layout.addWidget(QLabel("Resumen:"))
        self.snapshot_box = QTextEdit()
        self.snapshot_box.setReadOnly(True)
        self.snapshot_box.setMaximumHeight(90)
        self.snapshot_box.setPlainText("Sin datos cargados.")
        layout.addWidget(self.snapshot_box)

        self.setLayout(layout)

    # ── Carga de datos ────────────────────────────────────────────────────────

    def load_contacts(self):
        self.progress_bar.setVisible(True)
        self._loader = ContactsLoaderThread()
        self._loader.progress.connect(lambda _: None)
        self._loader.finished.connect(self._on_loaded)
        self._loader.error.connect(self._on_error)
        self._loader.start()
        self._load_labels()

    def _load_labels(self):
        self._lbl_thread = LabelsLoaderThread()
        self._lbl_thread.finished.connect(self._on_labels_loaded)
        self._lbl_thread.start()

    def _on_labels_loaded(self, labels: list):
        current = self.label_filter.currentText()
        self.label_filter.blockSignals(True)
        self.label_filter.clear()
        self.label_filter.addItem("Todas")
        for lbl in sorted(labels):
            self.label_filter.addItem(str(lbl))
        idx = self.label_filter.findText(current)
        self.label_filter.setCurrentIndex(idx if idx >= 0 else 0)
        self.label_filter.blockSignals(False)

    def _on_loaded(self, contacts: list):
        self.contacts = contacts
        self.has_loaded_once = True
        self._apply_filters()
        self.progress_bar.setVisible(False)
        self._refresh_snapshot(contacts)

    def _on_error(self, error: str):
        self.progress_bar.setVisible(False)
        QMessageBox.critical(self, "Error", f"Error cargando contactos:\n{error}")

    def ensure_loaded(self):
        if not self.has_loaded_once:
            self.load_contacts()

    # ── Filtrado ──────────────────────────────────────────────────────────────

    def _apply_filters(self):
        search = self.search_input.text().lower()
        label_sel = self.label_filter.currentText()
        filtered = []
        for c in self.contacts:
            tag = (c.get("etiqueta_cliente") or "").strip()
            if not tag:
                tags_list = c.get("intereses_tags", []) or c.get("intereses", []) or []
                tag = tags_list[0] if tags_list else ""
            if label_sel != "Todas" and tag != label_sel:
                continue
            if search:
                phone = str(c.get("whatsapp_number", "") or c.get("telefono", "")).lower()
                name = str(c.get("nombre", "")).lower()
                email = str(c.get("correo", "") or c.get("email", "")).lower()
                if search not in phone and search not in name and search not in email:
                    continue
            filtered.append(c)
        self.filtered_contacts = filtered
        self._fill_table(filtered)

    def _refresh_snapshot(self, contacts: list):
        total = len(contacts)
        labels: dict = {}
        for c in contacts:
            tag = (c.get("etiqueta_cliente") or "").strip()
            if not tag:
                tl = c.get("intereses_tags", []) or c.get("intereses", []) or []
                tag = tl[0] if tl else "sin etiqueta"
            labels[tag] = labels.get(tag, 0) + 1
        parts = [f"{k}: {v}" for k, v in sorted(labels.items())]
        self.snapshot_box.setPlainText(
            f"Total: {total}\nEtiquetas: {' | '.join(parts)}" if parts else f"Total: {total}"
        )

    # ── Tabla ─────────────────────────────────────────────────────────────────

    def _fill_table(self, contacts: list):
        self.table.setRowCount(len(contacts))
        for row, c in enumerate(contacts):
            chk = QTableWidgetItem()
            chk.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
            chk.setCheckState(Qt.CheckState.Unchecked)
            self.table.setItem(row, 0, chk)

            phone = c.get("telefono") or c.get("whatsapp_number") or "N/A"
            name = c.get("nombre", "N/A")
            email = c.get("correo") or c.get("email") or "N/A"
            tag = c.get("etiqueta_cliente") or ", ".join(c.get("intereses_tags", []) or []) or "N/A"
            date = c.get("actualizado_en") or c.get("fecha_registro") or "N/A"

            self.table.setItem(row, 1, QTableWidgetItem(str(phone)))
            self.table.setItem(row, 2, QTableWidgetItem(str(name)))
            self.table.setItem(row, 3, QTableWidgetItem(str(email)))
            self.table.setItem(row, 4, QTableWidgetItem(str(tag)))
            self.table.setItem(row, 5, QTableWidgetItem(str(date)))

        self.count_label.setText(f"Mostrando {len(contacts)} contacto(s)")

    def _get_checked_phones(self) -> list:
        phones = []
        for row in range(self.table.rowCount()):
            chk = self.table.item(row, 0)
            if chk and chk.checkState() == Qt.CheckState.Checked:
                item = self.table.item(row, 1)
                if item:
                    phones.append(item.text())
        return phones

    def _select_all(self):
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item:
                item.setCheckState(Qt.CheckState.Checked)

    def _deselect_all(self):
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item:
                item.setCheckState(Qt.CheckState.Unchecked)

    # ── Envío de mensajes ─────────────────────────────────────────────────────

    def _send_to_selected(self):
        phones = self._get_checked_phones()
        if not phones:
            QMessageBox.warning(self, "Aviso", "Marcá al menos un contacto (columna ✓)")
            return
        self._open_send_dialog(phones)

    def _send_to_all_visible(self):
        phones = [
            c.get("telefono") or c.get("whatsapp_number", "")
            for c in self.filtered_contacts
            if c.get("telefono") or c.get("whatsapp_number", "")
        ]
        if not phones:
            QMessageBox.warning(self, "Aviso", "No hay contactos visibles.")
            return
        reply = QMessageBox.question(
            self, "Confirmar",
            f"¿Enviar mensaje a los {len(phones)} contactos visibles?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._open_send_dialog(phones)

    def _open_send_dialog(self, phones: list):
        dlg = SendMessageDialog(phones, parent=self)
        if not dlg.exec():
            return
        message = dlg.get_message()
        if not message:
            QMessageBox.warning(self, "Aviso", "El mensaje no puede estar vacío.")
            return
        client = get_client()
        sent, errors = 0, []
        for phone in phones:
            ok, _, err = client.send_test_message(phone, message)
            if ok:
                sent += 1
            else:
                errors.append(f"{phone}: {err}")
        msg = f"✅ Enviados: {sent}/{len(phones)}"
        if errors:
            msg += f"\n⚠️ Errores:\n" + "\n".join(errors[:5])
            if len(errors) > 5:
                msg += f"\n… y {len(errors) - 5} más"
        QMessageBox.information(self, "Resultado envío", msg)

    # ── Edición de contactos ──────────────────────────────────────────────────

    def _on_double_click(self, index):
        self._open_edit_dialog(index.row())

    def _edit_from_button(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Aviso", "Seleccioná un contacto primero.")
            return
        self._open_edit_dialog(row)

    def _open_edit_dialog(self, row: int):
        if row < 0 or row >= len(self.filtered_contacts):
            return
        contact = self.filtered_contacts[row]
        dlg = ContactEditDialog(contact, parent=self)
        if not dlg.exec():
            return
        data = dlg.get_data()
        phone = contact.get("telefono") or contact.get("whatsapp_number", "")
        if not phone:
            QMessageBox.warning(self, "Aviso", "No se pudo obtener el teléfono del contacto.")
            return
        client = get_client()
        ok, _, err = client.update_contact(phone, data)
        if ok:
            QMessageBox.information(self, "Éxito", "Contacto actualizado.")
            self.load_contacts()
        else:
            QMessageBox.critical(self, "Error", f"No se pudo actualizar:\n{err}")

    # ── Eliminación ───────────────────────────────────────────────────────────

    def _delete_contact(self):
        row = self.table.currentRow()
        if row < 0 or row >= len(self.filtered_contacts):
            QMessageBox.warning(self, "Aviso", "Seleccioná un contacto primero.")
            return
        contact = self.filtered_contacts[row]
        phone = contact.get("telefono") or contact.get("whatsapp_number", "")
        name = contact.get("nombre", phone)
        reply = QMessageBox.question(
            self, "Confirmar eliminación",
            f"¿Eliminar el contacto '{name}' ({phone})?\nEsta acción no se puede deshacer.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        client = get_client()
        ok, _, err = client.delete_contact(phone)
        if ok:
            QMessageBox.information(self, "Éxito", f"Contacto {phone} eliminado.")
            self.load_contacts()
        else:
            QMessageBox.critical(self, "Error", f"No se pudo eliminar:\n{err}")

    # ── Exportación ───────────────────────────────────────────────────────────

    def _export_all(self):
        path, _ = QFileDialog.getSaveFileName(self, "Guardar", "", "Excel (*.xlsx)")
        if not path:
            return
        client = get_client()
        ok, data, err = client.export_all_contacts_xlsx()
        if ok:
            with open(path, "wb") as f:
                f.write(data)
            QMessageBox.information(self, "Éxito", f"Exportado a:\n{path}")
        else:
            QMessageBox.critical(self, "Error", f"Error exportando:\n{err}")

    def _export_by_label(self):
        client = get_client()
        ok, labels, err = client.get_labels()
        if not ok or not labels:
            QMessageBox.warning(self, "Aviso", f"No se pudieron obtener etiquetas:\n{err}")
            return
        label, accepted = QInputDialog.getItem(
            self, "Exportar por etiqueta", "Seleccioná la etiqueta:",
            [str(x) for x in labels], 0, False,
        )
        if not accepted:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar", f"contactos_{label}.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        ok, data, err = client.export_contacts_xlsx(label_filter=label)
        if ok:
            with open(path, "wb") as f:
                f.write(data)
            QMessageBox.information(self, "Éxito", f"Exportado por etiqueta '{label}'.")
        else:
            QMessageBox.critical(self, "Error", f"Error:\n{err}")

    def _export_by_date(self):
        date_range, accepted = QInputDialog.getText(
            self, "Exportar por fecha",
            "Rango (YYYY-MM-DD a YYYY-MM-DD) o un día (YYYY-MM-DD):",
        )
        if not accepted or not date_range.strip():
            return
        parts = [p.strip() for p in date_range.split(" a ") if p.strip()]
        if len(parts) == 1:
            date_from = date_to = parts[0]
        elif len(parts) == 2:
            date_from, date_to = parts
        else:
            QMessageBox.warning(self, "Aviso", "Formato inválido.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar", "contactos_fechas.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        client = get_client()
        ok, data, err = client.export_contacts_xlsx(date_from=date_from, date_to=date_to)
        if ok:
            with open(path, "wb") as f:
                f.write(data)
            QMessageBox.information(self, "Éxito", "Exportación por fecha completada.")
        else:
            QMessageBox.critical(self, "Error", f"Error:\n{err}")

    def _download_template(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar plantilla", "plantilla_contactos.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        client = get_client()
        ok, data, err = client.download_contacts_template()
        if ok:
            with open(path, "wb") as f:
                f.write(data)
            QMessageBox.information(self, "Éxito", "Plantilla descargada.")
        else:
            QMessageBox.critical(self, "Error", f"No se pudo descargar:\n{err}")

    # ── Importación ───────────────────────────────────────────────────────────

    def _import_contacts(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Abrir archivo", "",
            "Excel Files (*.xlsx);;CSV Files (*.csv)"
        )
        if not path:
            return
        reply = QMessageBox.question(
            self, "Confirmar importación",
            "¿Importar contactos desde el archivo seleccionado?\n"
            "Esto actualizará / creará contactos en la base de datos.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            headers = [
                str(cell.value or "").strip().lower()
                for cell in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            contacts = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                contact = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        contact[headers[i]] = str(val).strip() if val is not None else ""
                if contact:
                    contacts.append(contact)
            if not contacts:
                QMessageBox.warning(self, "Aviso", "No se encontraron datos en el archivo.")
                return
            client = get_client()
            ok, data, err = client.import_contacts(contacts)
            if ok:
                imported = (
                    data.get("importados", len(contacts))
                    if isinstance(data, dict) else len(contacts)
                )
                QMessageBox.information(self, "Éxito", f"✅ {imported} contactos importados.")
                self.load_contacts()
            else:
                QMessageBox.critical(self, "Error", f"Error importando:\n{err}")
        except ImportError:
            QMessageBox.critical(
                self, "Error",
                "Se requiere 'openpyxl' para leer Excel.\nInstalalo con: pip install openpyxl"
            )
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Error leyendo archivo:\n{str(exc)}")
