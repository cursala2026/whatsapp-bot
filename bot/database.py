"""bot/database.py — Capa de acceso a Firestore y utilidades de contactos.

Responsabilidades principales:
- Inicializar cliente Firestore.
- Upsert de perfiles comerciales.
- Idempotencia de mensajes (cache en memoria + persistencia en Firestore).
- Importacion/exportacion de contactos (CSV/XLSX).

No contiene logica de enrutamiento conversacional; eso vive en flow_user/flow_admin.
"""

import csv
import io
import os
import threading
import time as _time
from datetime import date, datetime
from typing import Any, Callable, Dict, List, Optional, Tuple

from fastapi import HTTPException

from bot.config import (
    logger,
    FIREBASE_CREDENTIALS_PATH,
    FIREBASE_PROJECT_ID,
    FIRESTORE_COLLECTION,
)
from bot.utils import (
    normalize_number,
    normalize_text_for_filter,
    normalize_interest_tag,
    build_contact_code,
    infer_argentina_province_from_phone,
    validate_bsuid,
    _normalize_intereses_backup,
)

try:
    import firebase_admin  # type: ignore[import-not-found]
    from firebase_admin import credentials, firestore  # type: ignore[import-not-found]
except Exception:
    firebase_admin = None
    credentials = None
    firestore = None

# ============================================================
# INICIALIZACIÓN DE FIRESTORE
# ============================================================

def init_firestore_client():
    if firebase_admin is None or credentials is None or firestore is None:
        logger.warning("firebase_admin no esta instalado. Firestore deshabilitado.")
        return None
    if firebase_admin._apps:
        return firestore.client()
    if not os.path.exists(FIREBASE_CREDENTIALS_PATH):
        logger.warning("No se encontro credencial de Firebase en: %s", FIREBASE_CREDENTIALS_PATH)
        logger.warning("Firestore deshabilitado hasta configurar FIREBASE_CREDENTIALS_PATH.")
        return None
    try:
        cred = credentials.Certificate(FIREBASE_CREDENTIALS_PATH)
        if FIREBASE_PROJECT_ID:
            firebase_admin.initialize_app(cred, {"projectId": FIREBASE_PROJECT_ID})
        else:
            firebase_admin.initialize_app(cred)
        logger.info("Firestore inicializado correctamente")
        return firestore.client()
    except Exception as e:
        logger.error("Error inicializando Firestore: %s", e)
        return None


# Singleton global
firestore_db = init_firestore_client()


# Etiquetas canónicas acordadas (no agregar etiquetas fuera de este set)
CANONICAL_CONTACT_LABELS = [
    "Redes y telecom",
    "END",
    "Diseño de cañdrias",
    "Soldadura",
    "Contacto IG",
    "diseño mecánico",
    "Pymes",
    "Instrumentación",
    "Cliente potencial",
    "Mineria",
    "LOGÍSTICA",
]


def _normalize_label_key(value: str) -> str:
    return normalize_interest_tag(str(value or "")).replace("_", " ").strip()


_LABEL_ALIAS_TO_CANONICAL = {
    # Redes
    "redes": "Redes y telecom",
    "redes y telecom": "Redes y telecom",
    "redes y telecomunicaciones": "Redes y telecom",
    "curso redes": "Redes y telecom",
    "curso telecom": "Redes y telecom",
    "redes y telecomunicaciones": "Redes y telecom",
    # END
    "end": "END",
    # Diseño de cañdrias
    "diseno de candrias": "Diseño de cañdrias",
    "diseno candrias": "Diseño de cañdrias",
    "candrias": "Diseño de cañdrias",
    # Soldadura
    "soldadura": "Soldadura",
    "curso soldadura": "Soldadura",
    # Contacto IG
    "contacto ig": "Contacto IG",
    "ig": "Contacto IG",
    # Diseño mecanico
    "diseno mecanico": "diseño mecánico",
    "diseño mecanico": "diseño mecánico",
    "diseno mecanica": "diseño mecánico",
    # Pymes
    "pymes": "Pymes",
    "capacitaciones empresas": "Pymes",
    "capacitaciones emp": "Pymes",
    "lead empresa": "Pymes",
    "interesado empresa": "Pymes",
    # Instrumentación
    "instrumentacion": "Instrumentación",
    "curso instrumentacion": "Instrumentación",
    # Cliente potencial
    "cliente potencial": "Cliente potencial",
    "interesado": "Cliente potencial",
    "interesado cursos": "Cliente potencial",
    "interesado profesional": "Cliente potencial",
    "interesado asesoria": "Cliente potencial",
    "lead profesional": "Cliente potencial",
    "lead asesoria empresa": "Cliente potencial",
    "lead asesoria persona": "Cliente potencial",
    # Mineria
    "mineria": "Mineria",
    "curso mineria": "Mineria",
    "exploracion minera": "Mineria",
    # Logistica
    "logistica": "LOGÍSTICA",
    "curso logistica": "LOGÍSTICA",
    "logistica y supply": "LOGÍSTICA",
}


def canonicalize_contact_label(raw_label: Optional[str]) -> str:
    """Devuelve etiqueta canónica. Si no coincide con el catálogo acordado, devuelve cadena vacía."""
    text = " ".join(str(raw_label or "").strip().split())
    if not text:
        return ""

    # Algunas fuentes guardan múltiples etiquetas separadas por '|'.
    parts = [" ".join(p.strip().split()) for p in text.split("|") if p and p.strip()]
    if not parts:
        parts = [text]

    canonical_by_norm = {_normalize_label_key(v): v for v in CANONICAL_CONTACT_LABELS}

    for part in parts:
        norm = _normalize_label_key(part)
        if norm in canonical_by_norm:
            return canonical_by_norm[norm]
        alias_hit = _LABEL_ALIAS_TO_CANONICAL.get(norm)
        if alias_hit:
            return alias_hit
    return ""


# ============================================================
# CACHE DE IDEMPOTENCIA EN MEMORIA
# ============================================================

_processed_msgs: dict = {}        # msg_id -> timestamp float
_processed_msgs_lock = threading.Lock()
_MSG_CACHE_TTL = 3600             # 1 hora: máx tiempo de reintento de Meta


def _run_bg(fn, *args, **kwargs) -> None:
    """Ejecuta fn en un hilo daemon (fire-and-forget). Para escrituras que no bloqueen la respuesta."""
    threading.Thread(target=fn, args=args, kwargs=kwargs, daemon=True).start()


# ============================================================
# UPSERT DE PERFIL
# ============================================================

def upsert_user_profile_firestore(
    whatsapp_number: str,
    nombre: Optional[str] = None,
    telefono: Optional[str] = None,
    intereses: Optional[list] = None,
    evento: str = "",
    extra_fields: Optional[dict] = None,
    etiqueta_cliente: Optional[str] = None,
    bsuid: Optional[str] = None,
):
    if firestore_db is None:
        return

    # Solo usar telefono real para la clave de contacto. Nunca derivar telefono desde BSUID.
    phone = telefono or ""
    normalized_phone = normalize_number(phone)
    if not normalized_phone:
        return

    provincia, area_code = infer_argentina_province_from_phone(normalized_phone)
    provincia_slug = normalize_interest_tag(provincia)

    payload = {
        "whatsapp_number": normalized_phone,
        "telefono": {
            "normalizado": normalized_phone,
            "e164": f"+{normalized_phone}",
            "codigo_area": area_code,
        },
        "provincia_por_numero": {
            "nombre": provincia,
            "slug": provincia_slug,
            "codigo_area": area_code,
        },
        "indicadores": {
            "tiene_telefono": True,
            "provincia_slug": provincia_slug,
        },
        "actualizado_en": firestore.SERVER_TIMESTAMP,
    }

    if bsuid:
        bsuid_validated = validate_bsuid(bsuid)
        if bsuid_validated:
            payload["bsuid"] = bsuid_validated
            payload["indicadores"]["tiene_bsuid"] = True

    # Regla de nombre:
    # - Si viene nombre explícito (agenda/import/admin), usarlo.
    # - Si NO viene y el contacto no está agendado, usar nombre por defecto de WhatsApp.
    provided_name = " ".join((nombre or "").strip().split())
    extra_nombre_contacto = ""
    extra_nombre_whatsapp = ""
    contacto_agendado = False
    if isinstance(extra_fields, dict):
        extra_nombre_contacto = " ".join(str(extra_fields.get("nombre_contacto", "")).strip().split())
        extra_nombre_whatsapp = " ".join(str(extra_fields.get("nombre_whatsapp", "")).strip().split())
        contacto_agendado = bool(extra_fields.get("contacto_agendado", False))

    # nombre_whatsapp NO sobreescribe el nombre del usuario: solo se usa como
    # campo de referencia en Firestore. El campo 'nombre' solo se actualiza si viene
    # de una fuente explicita (onboarding del bot, importacion de agenda, admin).
    resolved_name = provided_name or extra_nombre_contacto

    if resolved_name:
        clean_name = resolved_name
        payload["nombre"] = clean_name
        payload["nombre_normalizado"] = normalize_text_for_filter(clean_name)
        payload["indicadores"]["tiene_nombre"] = True

    if intereses:
        labels = [" ".join(str(item).strip().split()) for item in intereses if str(item).strip()]
        tags = [normalize_interest_tag(label) for label in labels]
        tags = [tag for tag in tags if tag]
        if labels:
            payload["intereses_labels"] = firestore.ArrayUnion(labels)
        if tags:
            payload["intereses_tags"] = firestore.ArrayUnion(tags)
            payload["indicadores_interes"] = {tag: True for tag in tags}
            primary_tag = tags[0]
            payload["etiqueta_interes"] = primary_tag.upper()
            payload["contacto_codigo"] = build_contact_code(normalized_phone, primary_tag)
            if not etiqueta_cliente:
                etiqueta_cliente = canonicalize_contact_label(primary_tag.upper())

    if "contacto_codigo" not in payload:
        payload["contacto_codigo"] = build_contact_code(normalized_phone)

    if extra_fields:
        extra_fields_clean = dict(extra_fields)
        if "etiqueta_cliente" in extra_fields_clean:
            canonical_extra = canonicalize_contact_label(extra_fields_clean.get("etiqueta_cliente"))
            if canonical_extra:
                extra_fields_clean["etiqueta_cliente"] = canonical_extra
            else:
                extra_fields_clean.pop("etiqueta_cliente", None)
        payload.update(extra_fields_clean)

    if etiqueta_cliente:
        canonical_label = canonicalize_contact_label(etiqueta_cliente)
        if canonical_label:
            payload["etiqueta_cliente"] = canonical_label

    try:
        firestore_db.collection(FIRESTORE_COLLECTION).document(normalized_phone).set(payload, merge=True)
    except Exception as e:
        logger.warning("Error guardando perfil en Firestore: %s", e)


def track_user_interest(
    whatsapp_number: str,
    interest_label: str,
    evento: str = "interes_detectado",
    etiqueta_cliente: Optional[str] = None,
):
    upsert_user_profile_firestore(
        whatsapp_number=whatsapp_number,
        telefono=whatsapp_number,
        intereses=[interest_label],
        evento=evento,
        etiqueta_cliente=etiqueta_cliente,
    )


# ============================================================
# IDEMPOTENCIA DE MENSAJES
# ============================================================

def _is_message_processed(msg_id: str) -> bool:
    """Chequea idempotencia. Primero memoria (O(1), sin red), luego Firestore como fallback post-reinicio."""
    if not msg_id:
        return False
    with _processed_msgs_lock:
        if msg_id in _processed_msgs:
            return True
    # Fallback a Firestore solo en caso de reinicio/instancia nueva
    if firestore_db is None:
        return False
    try:
        doc = firestore_db.collection("mensajes_procesados").document(msg_id).get()
        if doc.exists:
            with _processed_msgs_lock:
                _processed_msgs[msg_id] = _time.time()
        return doc.exists
    except Exception as e:
        logger.warning("[idempotency] Error consultando msg_id %s: %s", msg_id, e)
        return False


def _mark_message_processed(msg_id: str) -> None:
    """Marca el mensaje como procesado en memoria inmediatamente y persiste en Firestore en background."""
    if not msg_id:
        return
    now = _time.time()
    with _processed_msgs_lock:
        _processed_msgs[msg_id] = now
        # Limpiar entradas expiradas para evitar crecimiento ilimitado en RAM
        if len(_processed_msgs) > 5000:
            cutoff = now - _MSG_CACHE_TTL
            expired = [k for k, v in _processed_msgs.items() if v < cutoff]
            for k in expired:
                del _processed_msgs[k]
    # Persistir en Firestore en background (para durabilidad inter-instancias)
    _run_bg(_persist_msg_processed, msg_id)


def _persist_msg_processed(msg_id: str) -> None:
    if firestore_db is None:
        return
    try:
        firestore_db.collection("mensajes_procesados").document(msg_id).set({
            "procesado_en": firestore.SERVER_TIMESTAMP,
        })
    except Exception as e:
        logger.warning("[idempotency] Error persistiendo msg_id %s en Firestore: %s", msg_id, e)


# ============================================================
# IMPORTAR CONTACTOS DESDE BACKUP/CSV
# ============================================================

def import_contacts_backup_to_firestore(
    payload: dict,
    progress_callback: Optional[Callable[[int, int, int], None]] = None,
) -> dict:
    contactos = payload.get("contactos")
    if not isinstance(contactos, list):
        raise HTTPException(status_code=400, detail="Formato invalido: 'contactos' debe ser una lista")

    evento_default = " ".join(str(payload.get("evento_default", "importacion_backup")).strip().split()) or "importacion_backup"

    seen_phones = set()
    imported = 0
    skipped_no_phone = 0
    skipped_duplicates = 0
    skipped_existing = 0
    skipped_invalid = 0
    failed = 0
    failures = []
    total_contacts = len(contactos)
    
    summary = {
        "importados": 0,
        "omitidos_sin_telefono": 0,
        "omitidos_duplicados": 0,
        "omitidos_existentes": 0,
        "omitidos_invalidos": 0,
        "fallidos": 0,
    }

    def _tick_progress(processed: int) -> None:
        if not progress_callback:
            return
        percent = 100 if total_contacts == 0 else int((processed / total_contacts) * 100)
        progress_callback(processed, total_contacts, percent)

    for idx, item in enumerate(contactos, start=1):
        if not isinstance(item, dict):
            skipped_invalid += 1
            failures.append({"index": idx, "error": "item_no_es_objeto"})
            _tick_progress(idx)
            continue

        raw_phone = (
            item.get("whatsapp_number")
            or item.get("phone")
            or item.get("telefono")
            or item.get("numero")
        )
        normalized_phone = normalize_number(raw_phone)

        if not normalized_phone:
            skipped_no_phone += 1
            failures.append({"index": idx, "error": "telefono_invalido_o_ausente"})
            _tick_progress(idx)
            continue

        if normalized_phone in seen_phones:
            skipped_duplicates += 1
            _tick_progress(idx)
            continue
        seen_phones.add(normalized_phone)

        try:
            existing_doc = firestore_db.collection(FIRESTORE_COLLECTION).document(normalized_phone).get()
            if existing_doc.exists:
                skipped_existing += 1
                _tick_progress(idx)
                continue
        except Exception as e:
            failed += 1
            failures.append({"index": idx, "telefono": normalized_phone, "error": f"error_verificando_existencia: {e}"})
            _tick_progress(idx)
            continue

        nombre = " ".join(str(item.get("nombre", "")).strip().split())
        etiqueta_cliente = " ".join(str(item.get("etiqueta_cliente", "")).strip().split())
        intereses = _normalize_intereses_backup(item.get("intereses"))
        ultimo_evento = " ".join(str(item.get("ultimo_evento", evento_default)).strip().split()) or "importacion_backup"

        extra_fields: Dict[str, Any] = {}
        if isinstance(item.get("extra_fields"), dict):
            extra_fields.update(item.get("extra_fields", {}))

        extra_fields.setdefault("contacto_agendado", True)
        extra_fields.setdefault("agendado_por", "importacion_backup")

        try:
            upsert_user_profile_firestore(
                whatsapp_number=normalized_phone,
                nombre=nombre or None,
                telefono=normalized_phone,
                intereses=intereses or None,
                evento=ultimo_evento,
                extra_fields=extra_fields,
                etiqueta_cliente=etiqueta_cliente or None,
            )
            summary["importados"] += 1
        except Exception as e:
            summary["fallidos"] += 1
            failures.append({"index": idx, "telefono": normalized_phone, "error": str(e)})

# ============================================================
# CONSULTAS DE CONTACTOS PARA BROADCAST
# ============================================================

def get_all_distinct_tags_from_firestore(limit: int = 500) -> list:
    if firestore_db is None:
        return []
    try:
        docs = list(firestore_db.collection(FIRESTORE_COLLECTION).limit(limit).stream())
        tags = set()
        for doc in docs:
            data = doc.to_dict() or {}
            label = canonicalize_contact_label(data.get("etiqueta_cliente"))
            if label:
                tags.add(label)
        return sorted(tags)
    except Exception as e:
        logger.warning("Error obteniendo etiquetas de Firestore: %s", e)
        return []


def get_contacts_by_label(label: str, limit: int = 500) -> list:
    if firestore_db is None:
        return []
    try:
        canonical_target = canonicalize_contact_label(label)
        if not canonical_target:
            return []
        docs = list(firestore_db.collection(FIRESTORE_COLLECTION).limit(limit).stream())
        contacts = []
        for doc in docs:
            data = doc.to_dict() or {}
            current_label = canonicalize_contact_label(data.get("etiqueta_cliente"))
            if current_label != canonical_target:
                continue
            telefono = (data.get("telefono") or {}).get("normalizado") or doc.id
            if not telefono:
                continue
            nombre = " ".join(str(data.get("nombre", "")).strip().split())
            bsuid = data.get("bsuid") or ""
            contacts.append({"telefono": telefono, "nombre": nombre, "bsuid": bsuid})
        return contacts
    except Exception as e:
        logger.warning("Error consultando contactos por etiqueta '%s': %s", label, e)
        return []


def get_all_contacts_from_firestore(limit: int = 500) -> list:
    if firestore_db is None:
        return []
    try:
        docs = list(firestore_db.collection(FIRESTORE_COLLECTION).limit(limit).stream())
        contacts = []
        for doc in docs:
            data = doc.to_dict() or {}
            telefono = (data.get("telefono") or {}).get("normalizado") or doc.id
            if not telefono:
                continue
            nombre = " ".join(str(data.get("nombre", "")).strip().split())
            bsuid = data.get("bsuid") or ""
            contacts.append({"telefono": telefono, "nombre": nombre, "bsuid": bsuid})
        return contacts
    except Exception as e:
        logger.warning("Error obteniendo todos los contactos de Firestore: %s", e)
        return []


def build_contacts_saved_list_message(limit: int = 20) -> str:
    if firestore_db is None:
        return "⚠️ Firestore no configurado."

    safe_limit = max(1, min(limit, 50))
    docs = []

    try:
        if firestore is not None and hasattr(firestore, "Query"):
            docs = list(
                firestore_db
                .collection(FIRESTORE_COLLECTION)
                .order_by("actualizado_en", direction=firestore.Query.DESCENDING)
                .limit(safe_limit)
                .stream()
            )
        else:
            docs = list(
                firestore_db
                .collection(FIRESTORE_COLLECTION)
                .limit(safe_limit)
                .stream()
            )
    except Exception:
        docs = list(
            firestore_db
            .collection(FIRESTORE_COLLECTION)
            .limit(safe_limit)
            .stream()
        )

    if not docs:
        return "📭 No hay contactos guardados todavía."

    lines = [f"*CONTACTOS GUARDADOS* (últimos {len(docs)})", ""]
    for idx, doc in enumerate(docs, start=1):
        data = doc.to_dict() or {}
        nombre = " ".join(str(data.get("nombre", "")).strip().split()) or "(sin nombre)"
        telefono = (data.get("telefono", {}) or {}).get("normalizado") or doc.id
        etiqueta = canonicalize_contact_label(data.get("etiqueta_cliente")) or "sin_etiqueta"
        if len(nombre) > 28:
            nombre = nombre[:28].rstrip() + "..."
        lines.append(f"{idx}. {nombre} | {telefono} | {etiqueta}")

    lines.append("")
    lines.append("0. Volver al menú admin")
    return "\n".join(lines)


# ============================================================
# EXPORTACIÓN DE CONTACTOS A CSV
# ============================================================

def _parse_export_date(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    clean = " ".join(str(value).strip().split())
    if not clean:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(clean, fmt).date()
        except ValueError:
            continue
    return None


def _extract_activity_date(data: dict) -> Optional[date]:
    raw = data.get("actualizado_en") or data.get("ultima_actividad")
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if hasattr(raw, "date"):
        try:
            return raw.date()
        except Exception:
            pass
    if isinstance(raw, str):
        iso = raw.strip()[:10]
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(iso, fmt).date()
            except ValueError:
                continue
    return None


def _apply_export_filters(
    docs: List,
    label_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> List:
    canonical_label = canonicalize_contact_label(label_filter)
    from_date = _parse_export_date(date_from)
    to_date = _parse_export_date(date_to)

    if not canonical_label and not from_date and not to_date:
        return docs

    filtered: List = []
    for doc in docs:
        data = doc.to_dict() or {}

        if canonical_label:
            current_label = canonicalize_contact_label(data.get("etiqueta_cliente"))
            if current_label != canonical_label:
                continue

        if from_date or to_date:
            activity_date = _extract_activity_date(data)
            if activity_date is None:
                continue
            if from_date and activity_date < from_date:
                continue
            if to_date and activity_date > to_date:
                continue

        filtered.append(doc)

    return filtered


def get_contact_label_counts_from_firestore(limit: int = 5000) -> List[Tuple[str, int]]:
    """Devuelve conteo de contactos por etiqueta canónica (solo etiquetas permitidas)."""
    if firestore_db is None:
        return []
    counts = {label: 0 for label in CANONICAL_CONTACT_LABELS}
    try:
        docs = _filter_exportable_docs(_fetch_all_firestore_docs(limit=limit))
        for doc in docs:
            data = doc.to_dict() or {}
            label = canonicalize_contact_label(data.get("etiqueta_cliente"))
            if label in counts:
                counts[label] += 1
        return [(label, counts[label]) for label in CANONICAL_CONTACT_LABELS if counts[label] > 0]
    except Exception as e:
        logger.warning("Error obteniendo conteo por etiqueta: %s", e)
        return []


def export_all_contacts_to_csv_bytes(
    limit: int = 1000,
    label_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> Tuple[bytes, int]:
    """Exporta todos los contactos de Firestore como bytes CSV (UTF-8 BOM).

    Retorna (csv_bytes, cantidad_de_contactos).
    El formato es compatible con Excel y con el formato de la herramienta
    de recuperación de contactos Node.js.
    """
    if firestore_db is None:
        return b"", 0

    try:
        docs = list(firestore_db.collection(FIRESTORE_COLLECTION).limit(limit).stream())
    except Exception as e:
        logger.warning("export_all_contacts_to_csv_bytes: error leyendo Firestore: %s", e)
        return b"", 0

    docs = _filter_exportable_docs(docs)
    docs = _apply_export_filters(docs, label_filter=label_filter, date_from=date_from, date_to=date_to)

    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)

    writer.writerow([
        "Numero", "Nombre", "Etiqueta", "Provincia", "Intereses", "Ultima_Actividad",
    ])

    count = 0
    for doc in docs:
        data = doc.to_dict() or {}

        telefono_dict = data.get("telefono") or {}
        numero = (
            telefono_dict.get("e164")
            or telefono_dict.get("normalizado")
            or doc.id
        )

        nombre = " ".join(str(data.get("nombre") or data.get("nombre_whatsapp") or data.get("nombre_contacto") or "").strip().split())
        etiqueta = canonicalize_contact_label(data.get("etiqueta_cliente"))

        prov_dict = data.get("provincia_por_numero") or {}
        provincia = str(prov_dict.get("nombre", "")).strip()

        intereses_labels = data.get("intereses_labels") or []
        intereses = " | ".join(str(x) for x in intereses_labels) if isinstance(intereses_labels, list) else ""

        actualizado_en = data.get("actualizado_en")
        if hasattr(actualizado_en, "isoformat"):
            ultima_actividad = actualizado_en.isoformat()[:10]
        elif hasattr(actualizado_en, "strftime"):
            ultima_actividad = actualizado_en.strftime("%Y-%m-%d")
        else:
            ultima_actividad = ""

        writer.writerow([
            numero, nombre, etiqueta, provincia, intereses, ultima_actividad,
        ])
        count += 1

    csv_text = output.getvalue()
    # UTF-8 BOM para compatibilidad con Excel
    return b"\xef\xbb\xbf" + csv_text.encode("utf-8"), count


def _fetch_all_firestore_docs(limit: int = 5000, fields: Optional[List[str]] = None) -> list:
    """Obtiene todos los documentos de FIRESTORE_COLLECTION en lotes de 500.

    Usa paginación con start_after para evitar timeouts de gRPC con grandes
    colecciones. Si `fields` está especificado, solo trae esos campos (más rápido).
    Propaga la excepción si falla para que el caller la maneje.
    """
    BATCH_SIZE = 500
    all_docs: list = []
    last_doc = None

    while len(all_docs) < limit:
        remaining = min(BATCH_SIZE, limit - len(all_docs))
        query = firestore_db.collection(FIRESTORE_COLLECTION).limit(remaining)
        if fields:
            query = query.select(fields)
        if last_doc is not None:
            query = query.start_after(last_doc)
        batch = list(query.stream())
        if not batch:
            break
        all_docs.extend(batch)
        last_doc = batch[-1]
        if len(batch) < remaining:
            break  # No hay más docs

    logger.info("_fetch_all_firestore_docs: %d docs obtenidos de '%s'", len(all_docs), FIRESTORE_COLLECTION)
    return all_docs


def _looks_like_real_whatsapp_number(value: str) -> bool:
    digits = normalize_number(value)
    if not digits:
        return False
    # WhatsApp argentino real: 549 + 10 digitos locales.
    return digits.startswith("549") and len(digits) == 13


def _contact_export_score(doc) -> int:
    data = doc.to_dict() or {}
    telefono_dict = data.get("telefono") or {}
    numero = telefono_dict.get("normalizado") or data.get("whatsapp_number") or doc.id
    nombre = " ".join(str(data.get("nombre", "")).strip().split())
    contacto_agendado = bool(data.get("contacto_agendado", False))

    score = 0
    if _looks_like_real_whatsapp_number(numero):
        score += 100
    if str(numero).startswith("549"):
        score += 50
    if nombre:
        score += 20
    if contacto_agendado:
        score += 5
    if data.get("export_blocked") or data.get("migracion.export_blocked") or (data.get("migracion") or {}).get("export_blocked"):
        score -= 1000
    return score


def _filter_exportable_docs(docs: List) -> List:
    """Filtra docs para exportación sin ocultar contactos válidos por nombre repetido.

    Regla:
    - Excluir siempre registros marcados como export_blocked.
    - Incluir siempre contactos con etiqueta canónica (vienen de WA Business).
    - Incluir contactos agendados manualmente.
    - Para contactos sin etiqueta, exigir número de WhatsApp argentino válido.
    - Deduplicar solo por número final (no por nombre).
    """

    def _is_valid_for_export(doc, data: dict) -> bool:
        telefono_dict = data.get("telefono") or {}
        numero = telefono_dict.get("normalizado") or data.get("whatsapp_number") or doc.id
        return _looks_like_real_whatsapp_number(numero)

    selected: List = []
    seen_numbers = set()

    for doc in docs:
        data = doc.to_dict() or {}
        is_blocked = bool(
            data.get("export_blocked")
            or data.get("migracion.export_blocked")
            or (data.get("migracion") or {}).get("export_blocked")
        )
        label = canonicalize_contact_label(data.get("etiqueta_cliente"))
        if is_blocked and not label:
            continue

        if not _is_valid_for_export(doc, data):
            continue

        telefono_dict = data.get("telefono") or {}
        numero = normalize_number(telefono_dict.get("normalizado") or data.get("whatsapp_number") or doc.id)
        if not numero:
            continue
        if numero in seen_numbers:
            continue

        seen_numbers.add(numero)
        selected.append(doc)

    return selected


def get_contacts_for_export(
    limit: int = 5000,
    label_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Obtiene y filtra contactos de Firestore, devolviendo una lista de dicts limpios."""
    if firestore_db is None:
        raise RuntimeError("Firestore no está inicializado.")

    _XLSX_FIELDS = [
        "nombre", "nombre_whatsapp", "nombre_contacto", "etiqueta_cliente", "telefono",
        "provincia_por_numero", "intereses_labels", "actualizado_en",
    ]
    docs = _filter_exportable_docs(_fetch_all_firestore_docs(limit=limit, fields=_XLSX_FIELDS))
    docs = _apply_export_filters(docs, label_filter=label_filter, date_from=date_from, date_to=date_to)

    contacts_data = []
    for doc in docs:
        data = doc.to_dict() or {}
        telefono_dict = data.get("telefono") or {}
        numero = telefono_dict.get("e164") or telefono_dict.get("normalizado") or doc.id
        nombre = " ".join(str(data.get("nombre") or data.get("nombre_whatsapp") or data.get("nombre_contacto") or "").strip().split())
        etiqueta = canonicalize_contact_label(data.get("etiqueta_cliente"))
        prov_dict = data.get("provincia_por_numero") or {}
        provincia = str(prov_dict.get("nombre", "")).strip()
        intereses_labels = data.get("intereses_labels") or []
        intereses = " | ".join(str(x) for x in intereses_labels) if isinstance(intereses_labels, list) else ""
        actualizado_en = data.get("actualizado_en")
        ultima_actividad = ""
        if hasattr(actualizado_en, "isoformat"):
            ultima_actividad = actualizado_en.isoformat()[:10]
        elif hasattr(actualizado_en, "strftime"):
            ultima_actividad = actualizado_en.strftime("%Y-%m-%d")

        contacts_data.append({
            "numero": numero, "nombre": nombre, "etiqueta": etiqueta, "provincia": provincia,
            "intereses": intereses, "ultima_actividad": ultima_actividad,
        })
    return contacts_data


def export_all_contacts_to_xlsx_bytes(
    limit: int = 5000,
    label_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> Tuple[bytes, int]:
    """Exporta todos los contactos de Firestore como bytes XLSX (Excel).

    Retorna (xlsx_bytes, cantidad_de_contactos).
    Genera un archivo Excel formateado con encabezados en negrita y filtros
    automáticos, listo para abrir desde un celular o PC.
    Lanza excepción si Firestore no está disponible o falla la lectura.
    """
    if firestore_db is None:
        raise RuntimeError("Firestore no está inicializado en Cloud Run. Verificar credenciales.")
    
    contacts_data = get_contacts_for_export(
        limit=limit,
        label_filter=label_filter,
        date_from=date_from,
        date_to=date_to,
    )
    
    if not contacts_data:
        return b"", 0

    file_bytes = _create_xlsx_from_contacts(contacts_data)
    return file_bytes, len(contacts_data)

def _create_xlsx_from_contacts(contacts: List[Dict[str, Any]]) -> bytes:

    try:
        import openpyxl  # type: ignore[import-not-found]
        from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore[import-not-found]
    except ImportError:
        raise RuntimeError("openpyxl no está instalado en el servidor.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contactos"

    headers = ["Numero", "Nombre", "Etiqueta", "Provincia", "Intereses", "Ultima_Actividad"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F1"
    ws.row_dimensions[1].height = 20

    # Anchos de columna ajustados
    col_widths = [18, 28, 22, 20, 35, 18]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    for row_idx, contact_data in enumerate(contacts, start=2):
        row_data = [
            contact_data.get("numero"), contact_data.get("nombre"), contact_data.get("etiqueta"),
            contact_data.get("provincia"), contact_data.get("intereses"), contact_data.get("ultima_actividad"),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_contacts_template_excel() -> bytes:
    """Genera plantilla descargable para que admins carguen nuevos contactos.
    
    Columnas: Nombre, Numero, Correo, Etiqueta de Interés
    Provincia se calculará automáticamente.
    """
    try:
        import openpyxl  # type: ignore[import-not-found]
        from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore[import-not-found]
    except ImportError:
        logger.error("openpyxl no esta instalado para generar plantilla")
        raise ImportError("openpyxl no esta instalado")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contactos"

    # Headers con formato
    headers = ["Nombre", "Numero", "Correo", "Etiqueta de Interés"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Ancho de columnas
    ws.column_dimensions["A"].width = 25  # Nombre
    ws.column_dimensions["B"].width = 18  # Numero
    ws.column_dimensions["C"].width = 25  # Correo
    ws.column_dimensions["D"].width = 30  # Etiqueta

    # Rows de ejemplo (vacías pero con instrucciones)
    example_data = [
        ["Juan Pérez", "5492613151234", "juan@example.com", "Redes y telecom"],
        ["María García", "5491145678900", "maria@example.com", "Mineria"],
    ]

    for row_idx, row_data in enumerate(example_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Agregar instrucciones en hoja separada
    ws_instruc = wb.create_sheet("Instrucciones")
    instruc_text = (
        "PLANTILLA DE CONTACTOS CURSALA\n\n"
        "Instrucciones:\n"
        "1. Usá la hoja 'Contactos' para agregar tus contactos\n"
        "2. Campos requeridos:\n"
        "   - Nombre: nombre completo o razón social\n"
        "   - Numero: número de WhatsApp (sin espacios ni símbolos)\n"
        "   - Correo: dirección de email válida\n"
        "   - Etiqueta: etiqueta de interés (opcional)\n\n"
        "3. Provincia se calculará automáticamente por el número\n"
        "4. Al importar, los duplicados serán omitidos\n"
        "5. Envía el archivo completado al bot como documento\n\n"
        "Etiquetas recomendadas:\n"
        "- Redes y telecom\n"
        "- Mineria\n"
        "- Soldadura\n"
        "- Instrumentación\n"
        "- Pymes\n"
        "- Cliente potencial\n"
        "- Contacto IG\n"
    )

    ws_instruc["A1"] = instruc_text
    ws_instruc["A1"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws_instruc.column_dimensions["A"].width = 60
    for row in ws_instruc.iter_rows(min_row=1, max_row=20):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
